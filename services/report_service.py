#services/report_service.py

import os
from decimal import Decimal
from pathlib import Path
from typing import Optional

from services.general_ledger_generator import GeneralLedgerGenerator

# ---------------------------------------------------------
#   INPUT FILE LOCATIONS
# ---------------------------------------------------------
# Directory holding the source General Ledger workbook(s).
GL_INPUT_DIR = os.path.join("data", "sources", "input")
# Default General Ledger workbook used when no explicit file is supplied.
DEFAULT_GL_FILE = os.path.join(
    GL_INPUT_DIR, "LE General Ledger 1-1-26 thru 3-31-26.xlsx"
)
# Default reporting entity / company name.
DEFAULT_ENTITY = "LE"


def _resolve_gl_file(gl_file: Optional[str]) -> str:
    """
    Resolve the General Ledger workbook to use for the report.

    If an explicit path is given, it is used. Otherwise the default GL file is
    used, falling back to the first .xlsx workbook found in the input directory.
    """
    if gl_file:
        return gl_file

    if Path(DEFAULT_GL_FILE).exists():
        return DEFAULT_GL_FILE

    if os.path.isdir(GL_INPUT_DIR):
        for name in sorted(os.listdir(GL_INPUT_DIR)):
            if name.lower().endswith(".xlsx") and not name.startswith("~$"):
                return os.path.join(GL_INPUT_DIR, name)

    return DEFAULT_GL_FILE


def generate_quarterly_report_service():

    return [
        {
            "name": "Q1 Financial Report",
            "status": "pending"
        },
        {
            "name": "Q2 Financial Report",
            "status": "completed"
        }
    ]


def generate_annual_report_service(
    year: int = 2026,
    entity: str = DEFAULT_ENTITY,
    gl_file: Optional[str] = None,
):
    """
    Build an annual financial report (annual trial balance) for an entity by
    loading the required General Ledger workbook and aggregating every monthly
    period that falls within the requested year.

    Args:
        year: Reporting year to aggregate (e.g. 2026).
        entity: Company/entity name used to organize the ledger.
        gl_file: Optional path to the GL workbook. Defaults to the workbook in
            ``data/sources/input``.

    Returns:
        A single-item list containing the structured annual report.
    """
    resolved_file = _resolve_gl_file(gl_file)

    if not Path(resolved_file).exists():
        return [
            {
                "name": f"{year} Annual Financial Report",
                "status": "error",
                "error": f"General Ledger file not found: {resolved_file}",
            }
        ]

    # Load every GL entry from the workbook into the generator.
    generator = GeneralLedgerGenerator()
    generator.add_entries_from_file(entity, resolved_file)

    # Keep only the periods (yyyy-mm) that belong to the requested year.
    year_prefix = f"{year:04d}-"
    periods = [p for p in generator.get_periods(entity) if p.startswith(year_prefix)]

    # Aggregate each account across all periods of the year.
    annual_accounts: dict[str, dict] = {}
    for period in periods:
        for account_number, summary in generator.generate_trial_balance(entity, period).items():
            bucket = annual_accounts.setdefault(
                account_number,
                {"name": summary.name, "debit": Decimal("0"), "credit": Decimal("0")},
            )
            bucket["debit"] += summary.debit
            bucket["credit"] += summary.credit

    accounts = [
        {
            "account_number": account_number,
            "name": data["name"],
            "debit": str(data["debit"]),
            "credit": str(data["credit"]),
            "balance": str(data["debit"] - data["credit"]),
        }
        for account_number, data in sorted(annual_accounts.items())
    ]

    total_debit = sum((data["debit"] for data in annual_accounts.values()), Decimal("0"))
    total_credit = sum((data["credit"] for data in annual_accounts.values()), Decimal("0"))

    return [
        {
            "name": f"{year} Annual Financial Report",
            "status": "completed",
            "report_type": "Annual Trial Balance",
            "entity": entity,
            "source_file": resolved_file,
            "periods_covered": periods,
            "account_count": len(accounts),
            "totals": {
                "debit": str(total_debit),
                "credit": str(total_credit),
                "net": str(total_debit - total_credit),
            },
            "accounts": accounts,
        }
    ]




async def generate_excel_reconciliation_report(year: int, quarter: int) -> str:
    import openpyxl
    import tempfile
    import os
    import re
    from copy import copy
    from postgresql_db.database import get_pool
    from services.gl_persistence_service import get_consolidated_reconciliation

    base_data = await get_consolidated_reconciliation(year, quarter)

    pool = await get_pool()
    async with pool.acquire() as conn:
        bank_accounts = await conn.fetch(
            """
            SELECT ba.id as bank_account_id, ba.company_id, ba.account_number, b.name as bank_name
            FROM bank_account ba
            JOIN bank b ON ba.bank_id = b.id
            """
        )

        # Get GL accounts activity per month per company
        gl_activity_rows = await conn.fetch(
            """
            SELECT sf.company_id, a.account_number, a.account_name, EXTRACT(MONTH FROM e.entry_date) as m, sum(l.amount) as net_amount
            FROM gl_entry_lines l
            JOIN gl_entries e ON e.id = l.gl_entry_id
            JOIN gl_source_files sf ON sf.id = e.source_file_id
            LEFT JOIN chart_of_accounts_usa a ON a.id = l.account_id
            WHERE sf.status = 'saved' AND EXTRACT(YEAR FROM e.entry_date) = $1
            GROUP BY sf.company_id, a.account_number, a.account_name, EXTRACT(MONTH FROM e.entry_date)
            """, year
        )
        
        # Get Bank statement activity per month
        bank_stmt_rows = await conn.fetch(
            """
            SELECT account_id as bank_account_id, CAST(statement_month AS INTEGER) as m, beginning_balance, total_additions, total_subtractions
            FROM bank_statement
            WHERE statement_year = $1
            """, year
        )

    gl_activity_by_company = {}
    for r in gl_activity_rows:
        cid = r["company_id"]
        if cid not in gl_activity_by_company:
            gl_activity_by_company[cid] = {}
        
        raw_name = r['account_name'] or 'Unknown'
        clean_name = re.split(r'\s*Beginning Balance:', raw_name, flags=re.IGNORECASE)[0].strip()
        
        acct_key = f"{r['account_number'] or ''} - {clean_name}"
        if acct_key not in gl_activity_by_company[cid]:
            gl_activity_by_company[cid][acct_key] = {}
        
        m = int(r["m"]) if r["m"] else 0
        if m > 0:
            gl_activity_by_company[cid][acct_key][m] = float(r["net_amount"])

    bank_stmts_by_account = {}
    for r in bank_stmt_rows:
        bid = r["bank_account_id"]
        if bid not in bank_stmts_by_account:
            bank_stmts_by_account[bid] = {}
        m = r["m"]
        if m > 0:
            bank_stmts_by_account[bid][m] = {
                "beg": float(r["beginning_balance"]),
                "add": float(r["total_additions"]),
                "sub": float(r["total_subtractions"])
            }

    accounts_by_company = {}
    for r in bank_accounts:
        cid = r["company_id"]
        accounts_by_company.setdefault(cid, []).append(r)

    template_path = r"C:\Users\AlvinTsang\Desktop\Projects\internal_portal\data\sources\template\Trial_Balance_Template_new.xlsx"
    wb = openpyxl.load_workbook(template_path)
    
    if "Recon" not in wb.sheetnames:
        raise Exception("Template missing 'Recon' tab")
        
    source_ws = wb["Recon"]

    def find_row_by_label(ws, search_text, col=2, start_row=1):
        search_text = search_text.lower()
        for row in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val and search_text in str(val).lower():
                return row
        return None

    for company in base_data.get("companies", []):
        cid = company["company_id"]
        entity = company["company_name"]
        
        banks = accounts_by_company.get(cid, [])
        for bank in banks:
            bank_name = bank["bank_name"] or "Bank"
            acct_num = bank["account_number"] or "0000"
            last_4 = str(acct_num)[-4:]
            
            abbrev = str(entity).split()[0][:10]
            bank_short = bank_name[:3].upper()
            sheet_name = f"{abbrev} {bank_short}-{last_4}"[:31]
            
            target_ws = wb.copy_worksheet(source_ws)
            target_ws.title = sheet_name
            
            target_ws['B2'] = entity
            target_ws['D2'] = bank_name
            target_ws['G2'] = acct_num
            target_ws['B3'] = "Automated"
            target_ws['G3'] = year

            # --- POPULATE BOOK BALANCE SECTION ---
            comp_gl = gl_activity_by_company.get(cid, {})
            acct_keys = list(comp_gl.keys())
            
            start_activity_row = 11
            num_accounts = len(acct_keys)
            n_insert = max(0, num_accounts - 9)
            
            if n_insert > 0:
                target_ws.insert_rows(20, amount=n_insert)
                
                # Shift merged cells down
                merged_ranges = list(target_ws.merged_cells.ranges)
                target_ws.merged_cells.ranges.clear()
                for m_range in merged_ranges:
                    if m_range.min_row >= 20:
                        m_range.shift(0, n_insert)
                    target_ws.merged_cells.add(m_range)
                    
                from openpyxl.utils import get_column_letter
                
                # Copy formatting from row 19 (the last original activity row)
                for r_idx in range(20, 20 + n_insert):
                    for c_idx in range(1, 17):
                        source_cell = target_ws.cell(row=19, column=c_idx)
                        target_cell = target_ws.cell(row=r_idx, column=c_idx)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)

                # Repair the formulas shifted by insert_rows
                for c in range(4, 17):
                    col = get_column_letter(c)
                    target_ws.cell(row=20+n_insert, column=c).value = f"=SUM({col}11:{col}{19+n_insert})"
                    target_ws.cell(row=22+n_insert, column=c).value = f"={col}10+{col}{20+n_insert}"
                    target_ws.cell(row=29+n_insert, column=c).value = f"={col}{25+n_insert}+{col}{26+n_insert}+{col}{27+n_insert}"
                    target_ws.cell(row=39+n_insert, column=c).value = f"=SUM({col}{32+n_insert}:{col}{38+n_insert})"
                    target_ws.cell(row=42+n_insert, column=c).value = f"={col}{22+n_insert}+{col}{39+n_insert}"
                    target_ws.cell(row=43+n_insert, column=c).value = f"={col}{29+n_insert}"
                    if c <= 6:
                        target_ws.cell(row=45+n_insert, column=c).value = f"=ROUND({col}{42+n_insert}-{col}{43+n_insert},2)"
                    else:
                        target_ws.cell(row=45+n_insert, column=c).value = f"=IF(AND({col}{25+n_insert}=0,{col}{26+n_insert}=0,{col}{27+n_insert}=0),\"\",ROUND({col}{42+n_insert}-{col}{43+n_insert},2))"
                    if c >= 7:
                        prev_col = get_column_letter(c-1)
                        target_ws.cell(row=10, column=c).value = f"={prev_col}{22+n_insert}"
                        
                    # Fix row totals in Column P for shifted Bank Stmt and Reconciling Items
                    if c == 16:
                        target_ws.cell(row=25+n_insert, column=c).value = f"=SUM(D{25+n_insert}:O{25+n_insert})"
                        for r in [26, 27, 32, 33, 34, 35, 36, 37, 38]:
                            target_ws.cell(row=r+n_insert, column=c).value = f"=SUM(D{r+n_insert}:O{r+n_insert})"
                
            for i, ak in enumerate(acct_keys):
                r_idx = start_activity_row + i
                target_ws.cell(row=r_idx, column=1).value = "Activity"
                target_ws.cell(row=r_idx, column=2).value = ak
                target_ws.cell(row=r_idx, column=3).value = ""
                target_ws.cell(row=r_idx, column=16).value = f"=SUM(D{r_idx}:O{r_idx})"
                
                # Plug in month values (Cols D=4 to O=15)
                for m in range(1, 13):
                    val = comp_gl[ak].get(m, 0.0)
                    if val != 0.0:
                        target_ws.cell(row=r_idx, column=3 + m).value = val
                        
            # --- POPULATE BANK BALANCE SECTION ---
            # Re-find the rows after potential insertions
            beg_bank_row = find_row_by_label(target_ws, "Beginning Bank Balance")
            add_bank_row = find_row_by_label(target_ws, "Total Deposits / Credits")
            sub_bank_row = find_row_by_label(target_ws, "Total Withdrawals / Debits")
            
            bank_id = bank["bank_account_id"]
            stmts = bank_stmts_by_account.get(bank_id, {})
            
            for m in range(1, 13):
                col_idx = 3 + m
                stmt = stmts.get(m)
                if stmt:
                    if beg_bank_row: target_ws.cell(row=beg_bank_row, column=col_idx).value = stmt["beg"]
                    if add_bank_row: target_ws.cell(row=add_bank_row, column=col_idx).value = stmt["add"]
                    if sub_bank_row: target_ws.cell(row=sub_bank_row, column=col_idx).value = -abs(stmt["sub"]) if stmt["sub"] else 0.0

            # --- POPULATE RECONCILING ITEMS ---
            deps_transit_row = find_row_by_label(target_ws, "Deposits in Transit")
            out_checks_row = find_row_by_label(target_ws, "Outstanding Checks")
            bank_charges_row = find_row_by_label(target_ws, "Bank Charges Not Yet")
            interest_row = find_row_by_label(target_ws, "Interest Earned Not")

            in_books = company.get("in_books_not_in_bank", [])
            deps_by_month = {m: 0.0 for m in range(1, 13)}
            checks_by_month = {m: 0.0 for m in range(1, 13)}
            for t in in_books:
                m = int(t.get("date", "2026-01-01").split("-")[1]) if t.get("date") else 1
                if t["amount"] > 0:
                    deps_by_month[m] += t["amount"] # Enter as positive
                elif t["amount"] < 0:
                    checks_by_month[m] += t["amount"] # Enter as negative

            in_bank = company.get("in_bank_not_in_books", [])
            charges_by_month = {m: 0.0 for m in range(1, 13)}
            interest_by_month = {m: 0.0 for m in range(1, 13)}
            for t in in_bank:
                m = int(t.get("date", "2026-01-01").split("-")[1]) if t.get("date") else 1
                if t["amount"] < 0:
                    charges_by_month[m] += t["amount"] # Enter as negative
                elif t["amount"] > 0:
                    interest_by_month[m] += t["amount"] # Enter as positive

            for m in range(1, 13):
                col_idx = 3 + m
                if deps_transit_row and deps_by_month[m] != 0.0:
                    target_ws.cell(row=deps_transit_row, column=col_idx).value = deps_by_month[m]
                if out_checks_row and checks_by_month[m] != 0.0:
                    target_ws.cell(row=out_checks_row, column=col_idx).value = checks_by_month[m]
                if bank_charges_row and charges_by_month[m] != 0.0:
                    target_ws.cell(row=bank_charges_row, column=col_idx).value = charges_by_month[m]
                if interest_row and interest_by_month[m] != 0.0:
                    target_ws.cell(row=interest_row, column=col_idx).value = interest_by_month[m]

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    
    return path
